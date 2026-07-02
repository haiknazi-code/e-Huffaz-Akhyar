from fastapi import FastAPI, APIRouter, HTTPException, Header, Depends, Query
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import io
import csv
import time
import zipfile
import logging
import secrets
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
import httpx
from datetime import datetime, timezone, date
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

SHARED_PASSWORD = os.environ.get('SHARED_PASSWORD', 'akhyar2026')
SIMPLE_TOKEN = os.environ.get('SIMPLE_TOKEN', 'ehuffaz-akhyar-2026-token')
GUEST_TOKEN = os.environ.get('GUEST_TOKEN', 'ehuffaz-guest-2026-token')

# Halaqah config with Google Sheet GIDs
SHEET_BASE = "https://docs.google.com/spreadsheets/d/e/2PACX-1vQRx5ARNXspqaCv5U2ouQ37RsR9Pj6UvIq-arocAvrXwD3UOx2k2v10ziiAGdAR07wB-06APyH4hK71/pub?output=csv"

HALAQAHS = [
    {"id": "haikal", "name": "HALAQAH MUALIM HAIKAL", "teacher": "Mualim Haikal", "gid": "0"},
    {"id": "syafiq", "name": "HALAQAH MUALIM SYAFIQ", "teacher": "Mualim Syafiq", "gid": "1261321526"},
    {"id": "amir", "name": "HALAQAH MUALIM AMIR", "teacher": "Mualim Amir", "gid": "1863792579"},
    {"id": "atiqah", "name": "HALAQAH MUALIMAH ATIQAH", "teacher": "Mualimah Atiqah", "gid": "1140812216"},
    {"id": "nawwar", "name": "HALAQAH MUALIMAH NAWWAR", "teacher": "Mualimah Nawwar", "gid": "1880547878"},
]

HALAQAH_BY_ID = {h["id"]: h for h in HALAQAHS}

# Simple in-memory cache
_students_cache: Dict[str, Dict[str, Any]] = {}
CACHE_TTL = 300  # 5 minutes


app = FastAPI()
api_router = APIRouter(prefix="/api")


# ---------- Models ----------
class LoginRequest(BaseModel):
    password: str


class LoginResponse(BaseModel):
    token: str
    message: str


class BuletinCreate(BaseModel):
    tajuk: str
    memo: str
    penghantar: str
    tarikh: str  # YYYY-MM-DD


class Buletin(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    tajuk: str
    memo: str
    penghantar: str
    tarikh: str
    created_at: str


class StudentAdd(BaseModel):
    nama: str


class StudentProfileSave(BaseModel):
    student_id: str
    halaqah_id: str
    nama: str
    selected_juzuk_surah: List[str] = []


class Student(BaseModel):
    id: str
    nama: str
    halaqah_id: str
    source: str  # "sheet" or "db"


class TasmiCreate(BaseModel):
    halaqah_id: str
    student_id: str
    student_nama: str
    tarikh: str
    tarikh_akhir: Optional[str] = None
    is_weekly: Optional[bool] = False
    mod: str  # "hafazan_baru", "murajaah", "tilawah", "iqra", "amoktha_khatam"
    juzuk_surah: Optional[str] = None
    iqra_level: Optional[str] = None
    muka_surat: Optional[str] = None
    jumlah: Optional[str] = None
    jumlah_type: Optional[str] = None  # "baris" or "muka_surat"
    keputusan: Optional[str] = None  # "mumtaz", "jayyid", "daif", "gagal_hantar", "tidak_hadir"
    catatan: Optional[str] = None


class TasmiRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    halaqah_id: str
    student_id: str
    student_nama: str
    tarikh: str
    tarikh_akhir: Optional[str] = None
    is_weekly: Optional[bool] = False
    mod: str
    juzuk_surah: Optional[str] = None
    iqra_level: Optional[str] = None
    muka_surat: Optional[str] = None
    jumlah: Optional[str] = None
    jumlah_type: Optional[str] = None
    keputusan: Optional[str] = None
    catatan: Optional[str] = None
    created_at: str


# ---------- Auth ----------
def verify_token(authorization: Optional[str] = Header(None)):
    """Allows both teacher and guest tokens (read-only endpoints)."""
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Unauthorized")
    token = authorization.replace("Bearer ", "")
    if token == SIMPLE_TOKEN:
        return "teacher"
    if token == GUEST_TOKEN:
        return "guest"
    raise HTTPException(status_code=401, detail="Invalid token")


def verify_teacher(authorization: Optional[str] = Header(None)):
    """Allows only the teacher token (write endpoints)."""
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Unauthorized")
    token = authorization.replace("Bearer ", "")
    if token != SIMPLE_TOKEN:
        raise HTTPException(status_code=403, detail="Tetamu tidak boleh melakukan tindakan ini")
    return "teacher"


@api_router.post("/auth/login", response_model=LoginResponse)
async def login(payload: LoginRequest):
    if payload.password != SHARED_PASSWORD:
        raise HTTPException(status_code=401, detail="Kata laluan salah")
    return LoginResponse(token=SIMPLE_TOKEN, message="Berjaya log masuk")


@api_router.post("/auth/guest", response_model=LoginResponse)
async def login_guest():
    return LoginResponse(token=GUEST_TOKEN, message="Berjaya log masuk sebagai tetamu")


# ---------- Halaqahs / Students ----------
async def fetch_students_from_sheet(halaqah_id: str) -> List[Dict[str, Any]]:
    now = time.time()
    cached = _students_cache.get(halaqah_id)
    if cached and now - cached["ts"] < CACHE_TTL:
        return cached["data"]

    halaqah = HALAQAH_BY_ID.get(halaqah_id)
    if not halaqah:
        return []

    url = f"{SHEET_BASE}&gid={halaqah['gid']}"
    students: List[Dict[str, Any]] = []
    import hashlib
    try:
        async with httpx.AsyncClient(timeout=15.0, follow_redirects=True) as hc:
            resp = await hc.get(url)
            resp.raise_for_status()
            text = resp.text
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        # Find name column heuristically
        for row in rows:
            if not row:
                continue
            for cell in row:
                cell_clean = (cell or "").strip()
                if not cell_clean:
                    continue
                lower = cell_clean.lower()
                # Skip header-like cells
                if lower in ("nama", "name", "bil", "no", "no.", "#", "pelajar", "nama pelajar", "senarai pelajar", "nama murid", "kelas tingkatan", "kelas", "tingkatan", "halaqah"):
                    continue
                if len(cell_clean) < 3:
                    continue
                # Must contain at least a letter
                if not any(c.isalpha() for c in cell_clean):
                    continue
                # Stable hash-based ID (immune to row reordering)
                name_hash = hashlib.md5(cell_clean.upper().encode("utf-8")).hexdigest()[:8]
                students.append({
                    "id": f"sheet_{halaqah_id}_{name_hash}",
                    "nama": cell_clean,
                    "halaqah_id": halaqah_id,
                    "source": "sheet"
                })
                break  # Take only first valid cell per row
    except Exception as e:
        logging.error(f"Failed to fetch students for {halaqah_id}: {e}")
        students = []

    _students_cache[halaqah_id] = {"ts": now, "data": students}
    return students


@api_router.get("/halaqahs")
async def get_halaqahs():
    return HALAQAHS


@api_router.get("/halaqahs/{halaqah_id}/students")
async def get_students(halaqah_id: str, _auth: bool = Depends(verify_token)):
    if halaqah_id not in HALAQAH_BY_ID:
        raise HTTPException(status_code=404, detail="Halaqah not found")

    sheet_students = await fetch_students_from_sheet(halaqah_id)
    db_students_cursor = db.extra_students.find({"halaqah_id": halaqah_id}, {"_id": 0})
    db_students = await db_students_cursor.to_list(1000)
    return sheet_students + db_students


@api_router.post("/halaqahs/{halaqah_id}/students")
async def add_student(halaqah_id: str, payload: StudentAdd, _auth: bool = Depends(verify_teacher)):
    if halaqah_id not in HALAQAH_BY_ID:
        raise HTTPException(status_code=404, detail="Halaqah not found")
    doc = {
        "id": f"db_{uuid.uuid4().hex[:8]}",
        "nama": payload.nama.strip(),
        "halaqah_id": halaqah_id,
        "source": "db"
    }
    await db.extra_students.insert_one(doc.copy())
    return doc


# ---------- Buletin ----------
@api_router.get("/buletin")
async def list_buletin(_auth: bool = Depends(verify_token)):
    items = await db.buletin.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return items


@api_router.post("/buletin", response_model=Buletin)
async def create_buletin(payload: BuletinCreate, _auth: bool = Depends(verify_teacher)):
    doc = {
        "id": str(uuid.uuid4()),
        "tajuk": payload.tajuk,
        "memo": payload.memo,
        "penghantar": payload.penghantar,
        "tarikh": payload.tarikh,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.buletin.insert_one(doc.copy())
    return doc


@api_router.delete("/buletin/{buletin_id}")
async def delete_buletin(buletin_id: str, _auth: bool = Depends(verify_teacher)):
    res = await db.buletin.delete_one({"id": buletin_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Buletin not found")
    return {"ok": True}


# ---------- Tasmi ----------
@api_router.post("/tasmi", response_model=TasmiRecord)
async def create_tasmi(payload: TasmiCreate, _auth: bool = Depends(verify_teacher)):
    doc = {
        "id": str(uuid.uuid4()),
        **payload.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.tasmi_records.insert_one(doc.copy())
    return doc


@api_router.get("/tasmi")
async def list_tasmi(
    halaqah_id: Optional[str] = None,
    student_id: Optional[str] = None,
    mod: Optional[str] = None,
    month: Optional[str] = None,  # YYYY-MM
    _auth: bool = Depends(verify_token)
):
    q: Dict[str, Any] = {}
    if halaqah_id:
        q["halaqah_id"] = halaqah_id
    if student_id:
        q["student_id"] = student_id
    if mod:
        q["mod"] = mod
    if month:
        q["tarikh"] = {"$regex": f"^{month}"}
    items = await db.tasmi_records.find(q, {"_id": 0}).sort("tarikh", -1).to_list(5000)
    return items


@api_router.delete("/tasmi/{record_id}")
async def delete_tasmi(record_id: str, _auth: bool = Depends(verify_teacher)):
    res = await db.tasmi_records.delete_one({"id": record_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Record not found")
    return {"ok": True}


# ---------- Analytics ----------
@api_router.get("/analytics/pending-halaqahs")
async def pending_halaqahs(_auth: bool = Depends(verify_token)):
    today = date.today().isoformat()
    submitted_ids = await db.tasmi_records.distinct("halaqah_id", {"tarikh": today})
    pending = [h for h in HALAQAHS if h["id"] not in submitted_ids]
    return {
        "tarikh": today,
        "pending": pending,
        "jumlah_belum": len(pending),
        "jumlah_total": len(HALAQAHS)
    }


def _safe_int(v):
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError, AttributeError):
        return 0


def _is_countable(r):
    """A record contributes to analytics if it has keputusan in (mumtaz/jayyid/daif) OR is_weekly."""
    if r.get("is_weekly"):
        return True
    return r.get("keputusan") in ("mumtaz", "jayyid", "daif")


def _date_range(start_iso, end_iso):
    """Return list of ISO date strings from start to end inclusive."""
    from datetime import date as _date
    try:
        y1, m1, d1 = [int(x) for x in start_iso.split("-")]
        y2, m2, d2 = [int(x) for x in end_iso.split("-")]
        d_start = _date(y1, m1, d1)
        d_end = _date(y2, m2, d2)
        if d_end < d_start:
            return [start_iso]
        days = []
        cur = d_start
        while cur <= d_end:
            days.append(cur.isoformat())
            cur = _date.fromordinal(cur.toordinal() + 1)
        return days
    except Exception:
        return [start_iso]


@api_router.get("/analytics/top-students")
async def top_students(month: Optional[str] = None, _auth: bool = Depends(verify_token)):
    q: Dict[str, Any] = {}
    if month:
        q["tarikh"] = {"$regex": f"^{month}"}
    records = await db.tasmi_records.find(q, {"_id": 0}).to_list(10000)
    
    hafazan_count: Dict[str, Dict[str, Any]] = {}
    murajaah_count: Dict[str, Dict[str, Any]] = {}
    for r in records:
        if not _is_countable(r):
            continue
        sid = r["student_id"]
        nama = r["student_nama"]
        hid = r["halaqah_id"]
        key = f"{sid}|{hid}"
        if r["mod"] == "hafazan_baru":
            entry = hafazan_count.setdefault(key, {"student_id": sid, "nama": nama, "halaqah_id": hid, "total_baris": 0, "count": 0})
            entry["total_baris"] += _safe_int(r.get("jumlah"))
            entry["count"] += 1
        if r["mod"] == "murajaah":
            entry = murajaah_count.setdefault(key, {"student_id": sid, "nama": nama, "halaqah_id": hid, "total_muka_surat": 0, "count": 0})
            entry["total_muka_surat"] += _safe_int(r.get("jumlah"))
            entry["count"] += 1

    top_hafazan = sorted(hafazan_count.values(), key=lambda x: -x["total_baris"])[:10]
    top_murajaah = sorted(murajaah_count.values(), key=lambda x: -x["total_muka_surat"])[:10]
    return {"hafazan_baru": top_hafazan, "murajaah": top_murajaah}


@api_router.get("/analytics/top-halaqahs")
async def top_halaqahs(month: Optional[str] = None, _auth: bool = Depends(verify_token)):
    q: Dict[str, Any] = {}
    if month:
        q["tarikh"] = {"$regex": f"^{month}"}
    records = await db.tasmi_records.find(q, {"_id": 0}).to_list(10000)
    stats: Dict[str, Dict[str, Any]] = {}
    for h in HALAQAHS:
        stats[h["id"]] = {
            "halaqah_id": h["id"],
            "name": h["name"],
            "hafazan_baru": 0, "murajaah": 0, "tilawah": 0, "iqra": 0, "amoktha_khatam": 0,
            "hafazan_baru_baris": 0, "murajaah_muka_surat": 0,
            "mumtaz": 0, "jayyid": 0, "daif": 0,
            "total": 0
        }
    for r in records:
        hid = r["halaqah_id"]
        if hid not in stats:
            continue
        mod = r["mod"]
        if mod in stats[hid]:
            stats[hid][mod] += 1
        keputusan = r.get("keputusan")
        countable = _is_countable(r)
        if countable and mod == "hafazan_baru":
            stats[hid]["hafazan_baru_baris"] += _safe_int(r.get("jumlah"))
        if countable and mod == "murajaah":
            stats[hid]["murajaah_muka_surat"] += _safe_int(r.get("jumlah"))
        if keputusan in ("mumtaz", "jayyid", "daif"):
            stats[hid][keputusan] += 1
        stats[hid]["total"] += 1
    ranking = sorted(stats.values(), key=lambda x: -(x["hafazan_baru_baris"] + x["murajaah_muka_surat"]))
    return ranking


@api_router.get("/analytics/overall")
async def overall(month: Optional[str] = None, _auth: bool = Depends(verify_token)):
    q: Dict[str, Any] = {}
    if month:
        q["tarikh"] = {"$regex": f"^{month}"}
    records = await db.tasmi_records.find(q, {"_id": 0}).to_list(10000)
    hantar = 0
    tidak_hantar = 0
    tidak_hadir = 0
    for r in records:
        kep = r.get("keputusan")
        if kep == "gagal_hantar":
            tidak_hantar += 1
        elif kep == "tidak_hadir":
            tidak_hadir += 1
        else:
            hantar += 1
    total = hantar + tidak_hantar + tidak_hadir
    return {
        "hantar": hantar,
        "tidak_hantar": tidak_hantar,
        "tidak_hadir": tidak_hadir,
        "total": total
    }


@api_router.get("/analytics/weak-students")
async def weak_students(month: Optional[str] = None, _auth: bool = Depends(verify_token)):
    q: Dict[str, Any] = {}
    if month:
        q["tarikh"] = {"$regex": f"^{month}"}
    records = await db.tasmi_records.find(q, {"_id": 0}).to_list(10000)
    by_student: Dict[str, Dict[str, Any]] = {}
    for r in records:
        sid = r["student_id"]
        key = f"{sid}|{r['halaqah_id']}"
        entry = by_student.setdefault(key, {
            "student_id": sid, "nama": r["student_nama"], "halaqah_id": r["halaqah_id"],
            "gagal_hantar": 0, "tidak_hadir": 0, "daif": 0, "total": 0
        })
        kep = r.get("keputusan")
        if kep == "gagal_hantar":
            entry["gagal_hantar"] += 1
        elif kep == "tidak_hadir":
            entry["tidak_hadir"] += 1
        elif kep == "daif":
            entry["daif"] += 1
        entry["total"] += 1
    items = list(by_student.values())
    return {
        "kerap_gagal": sorted(items, key=lambda x: -x["gagal_hantar"])[:10],
        "kerap_tidak_hadir": sorted(items, key=lambda x: -x["tidak_hadir"])[:10],
        "kerap_daif": sorted(items, key=lambda x: -x["daif"])[:10],
    }


@api_router.get("/analytics/halaqah-report")
async def halaqah_report(halaqah_id: str, month: str, _auth: bool = Depends(verify_token)):
    q = {"halaqah_id": halaqah_id, "tarikh": {"$regex": f"^{month}"}}
    records = await db.tasmi_records.find(q, {"_id": 0}).to_list(10000)
    summary = {
        "hafazan_baru": 0, "murajaah": 0, "tilawah": 0, "iqra": 0, "amoktha_khatam": 0,
        "mumtaz": 0, "jayyid": 0, "daif": 0, "gagal_hantar": 0, "tidak_hadir": 0,
        "total": len(records)
    }
    for r in records:
        if r["mod"] in summary:
            summary[r["mod"]] += 1
        kep = r.get("keputusan")
        if kep in summary:
            summary[kep] += 1
    return {"halaqah_id": halaqah_id, "month": month, "summary": summary, "records": records}


@api_router.get("/analytics/jejak-hafazan")
async def jejak_hafazan(_auth: bool = Depends(verify_token)):
    """Road map menunjukkan kedudukan terkini setiap pelajar berdasarkan hafazan_baru terakhir."""
    today_iso = date.today().isoformat()

    # Stages in display order
    stages_order = [
        "IQRA",
        "JUZ 30", "JUZ 29", "JUZ 28",
        "AS-SAJDAH", "AR-RAHMAN", "AL-WAQIAH", "YASIN",
    ] + [f"JUZ {i}" for i in range(1, 28)] + ["TAMAT"]

    def map_to_stage(juzuk_surah: Optional[str], mod: Optional[str], iqra_level: Optional[str]) -> Optional[str]:
        if mod == "iqra" or (iqra_level and iqra_level.upper().startswith("IQRA")):
            return "IQRA"
        if not juzuk_surah:
            return None
        js = juzuk_surah.strip()
        js_upper = js.upper()
        # Surah handling
        if "AS-SAJDAH" in js_upper or "SAJDAH" in js_upper:
            return "AS-SAJDAH"
        if "AR-RAHMAN" in js_upper or "RAHMAN" in js_upper:
            return "AR-RAHMAN"
        if "AL-WAQIAH" in js_upper or "WAQIAH" in js_upper:
            return "AL-WAQIAH"
        if "YASIN" in js_upper or "YASSIN" in js_upper:
            return "YASIN"
        # Juzuk N
        if js_upper.startswith("JUZUK") or js_upper.startswith("JUZ"):
            digits = "".join(ch for ch in js if ch.isdigit())
            if digits:
                n = int(digits)
                if 1 <= n <= 30:
                    return f"JUZ {n}"
        return None

    # Gather all students
    all_students = []
    for h in HALAQAHS:
        sheet_students = await fetch_students_from_sheet(h["id"])
        db_students = await db.extra_students.find({"halaqah_id": h["id"]}, {"_id": 0}).to_list(1000)
        for s in sheet_students + db_students:
            all_students.append({
                "student_id": s["id"],
                "halaqah_id": h["id"],
                "halaqah_nama": h["name"],
                "nama": s["nama"],
            })

    # For each student, find their latest hafazan record (hafazan_baru OR iqra) up to today
    student_stage: Dict[str, Dict[str, Any]] = {}
    for s in all_students:
        latest = await db.tasmi_records.find_one(
            {
                "student_id": s["student_id"],
                "halaqah_id": s["halaqah_id"],
                "mod": {"$in": ["hafazan_baru", "iqra"]},
                "tarikh": {"$lte": today_iso},
            },
            {"_id": 0},
            sort=[("tarikh", -1), ("created_at", -1)]
        )
        if not latest:
            continue
        stage = map_to_stage(latest.get("juzuk_surah"), latest.get("mod"), latest.get("iqra_level"))
        if not stage:
            continue
        student_stage[f'{s["student_id"]}|{s["halaqah_id"]}'] = {
            **s,
            "stage": stage,
            "tarikh": latest.get("tarikh"),
            "muka_surat": latest.get("muka_surat"),
            "iqra_level": latest.get("iqra_level"),
        }

    # Group by stage
    stage_labels = {"IQRA": "MULA"}
    stages_data = []
    for stage in stages_order:
        members = [v for v in student_stage.values() if v["stage"] == stage]
        members.sort(key=lambda x: x["nama"] or "")
        stages_data.append({
            "stage": stage,
            "label": stage_labels.get(stage, stage),
            "students": members,
            "count": len(members),
        })

    return {"stages": stages_data, "total_pelajar_tracked": len(student_stage)}


@api_router.get("/analytics/kemajuan-hafazan")
async def kemajuan_hafazan(_auth: bool = Depends(verify_token)):
    """Ranking semua pelajar dari semua halaqah berdasarkan jumlah juzuk hafazan."""
    profiles = await db.student_profiles.find({}, {"_id": 0}).to_list(10000)
    items = []
    for p in profiles:
        selected = p.get("selected_juzuk_surah", []) or []
        jumlah_juzuk = sum(1 for x in selected if x.startswith("Juzuk"))
        jumlah_surah = sum(1 for x in selected if not x.startswith("Juzuk"))
        if jumlah_juzuk == 0 and jumlah_surah == 0:
            continue
        items.append({
            "student_id": p.get("student_id"),
            "halaqah_id": p.get("halaqah_id"),
            "halaqah_nama": HALAQAH_NAME_BY_ID.get(p.get("halaqah_id"), p.get("halaqah_id")),
            "nama": p.get("nama"),
            "jumlah_juzuk": jumlah_juzuk,
            "jumlah_surah": jumlah_surah,
            "total_skor": jumlah_juzuk + (jumlah_surah * 0.1),  # juzuk primary, surah tie-break
        })
    items.sort(key=lambda x: (-x["jumlah_juzuk"], -x["jumlah_surah"], x["nama"] or ""))
    return items


@api_router.get("/analytics/student-performance")
async def student_performance(halaqah_id: str, month: Optional[str] = None, _auth: bool = Depends(verify_token)):
    """Returns per-student daily aggregates: sum of baris (hafazan_baru) and muka surat (murajaah).
    Weekly records (is_weekly=True) are distributed evenly across the date range.
    Optional 'month' filter (YYYY-MM): only show data within that month."""
    q: Dict[str, Any] = {"halaqah_id": halaqah_id}
    records = await db.tasmi_records.find(q, {"_id": 0}).to_list(10000)
    by_student: Dict[str, Dict[str, Any]] = {}
    for r in records:
        sid = r["student_id"]
        entry = by_student.setdefault(sid, {
            "student_id": sid, "nama": r["student_nama"], "halaqah_id": halaqah_id,
            "total_baris": 0, "total_muka_surat": 0,
            "daily": {}
        })

        is_weekly = bool(r.get("is_weekly"))
        countable = _is_countable(r)
        jumlah_val = _safe_int(r.get("jumlah"))

        if is_weekly and r.get("tarikh_akhir"):
            days = _date_range(r["tarikh"], r["tarikh_akhir"])
        else:
            days = [r["tarikh"]]

        per_day_amount = (jumlah_val / len(days)) if (countable and days) else 0

        for tarikh in days:
            # Apply month filter at day level so weekly records spanning months are split correctly
            if month and not tarikh.startswith(month):
                continue
            d = entry["daily"].setdefault(tarikh, {"tarikh": tarikh, "hafazan_baru": 0, "murajaah": 0, "total": 0})
            if countable and r["mod"] == "hafazan_baru":
                d["hafazan_baru"] += per_day_amount
                entry["total_baris"] += per_day_amount
            if countable and r["mod"] == "murajaah":
                d["murajaah"] += per_day_amount
                entry["total_muka_surat"] += per_day_amount
            d["total"] += 1

    result = []
    for s in by_student.values():
        # Round daily values for cleanliness
        for d in s["daily"].values():
            d["hafazan_baru"] = round(d["hafazan_baru"], 1)
            d["murajaah"] = round(d["murajaah"], 1)
        s["total_baris"] = round(s["total_baris"], 1)
        s["total_muka_surat"] = round(s["total_muka_surat"], 1)
        s["daily"] = sorted(s["daily"].values(), key=lambda x: x["tarikh"])
        result.append(s)
    return sorted(result, key=lambda x: x["nama"])


# ---------- Student Profile ----------
@api_router.post("/student-profile")
async def save_student_profile(payload: StudentProfileSave, _auth: bool = Depends(verify_teacher)):
    key = {"student_id": payload.student_id, "halaqah_id": payload.halaqah_id}
    doc = {
        **key,
        "nama": payload.nama,
        "selected_juzuk_surah": payload.selected_juzuk_surah,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.student_profiles.update_one(key, {"$set": doc}, upsert=True)
    return doc


@api_router.get("/student-profile/halaqah/{halaqah_id}")
async def get_halaqah_profiles_summary(halaqah_id: str, _auth: bool = Depends(verify_token)):
    """Returns summary (jumlah_hafazan, hafazan_terkini, rankings) for ALL students in a halaqah."""
    if halaqah_id not in HALAQAH_BY_ID:
        raise HTTPException(status_code=404, detail="Halaqah not found")

    today_iso = date.today().isoformat()

    # 1. Build student list (sheet + db extras)
    sheet_students = await fetch_students_from_sheet(halaqah_id)
    db_students = await db.extra_students.find({"halaqah_id": halaqah_id}, {"_id": 0}).to_list(1000)
    all_students = sheet_students + db_students

    # 2. Fetch all profiles for this halaqah (juzuk selections)
    profile_cursor = db.student_profiles.find({"halaqah_id": halaqah_id}, {"_id": 0})
    profiles = await profile_cursor.to_list(5000)
    profile_map = {p["student_id"]: p for p in profiles}

    # 3. Fetch all tasmi records (for latest hafazan + rankings)
    records = await db.tasmi_records.find({}, {"_id": 0}).to_list(20000)

    # Compute global rankings
    hafazan_count: Dict[str, Dict[str, Any]] = {}
    murajaah_count: Dict[str, Dict[str, Any]] = {}
    for r in records:
        if not _is_countable(r):
            continue
        sid = r["student_id"]
        hid = r["halaqah_id"]
        k = f"{sid}|{hid}"
        if r["mod"] == "hafazan_baru":
            e = hafazan_count.setdefault(k, {"student_id": sid, "halaqah_id": hid, "total_baris": 0})
            e["total_baris"] += _safe_int(r.get("jumlah"))
        if r["mod"] == "murajaah":
            e = murajaah_count.setdefault(k, {"student_id": sid, "halaqah_id": hid, "total_muka_surat": 0})
            e["total_muka_surat"] += _safe_int(r.get("jumlah"))

    sorted_hafazan = sorted(hafazan_count.values(), key=lambda x: -x["total_baris"])
    sorted_murajaah = sorted(murajaah_count.values(), key=lambda x: -x["total_muka_surat"])
    rank_hafazan_map = {f'{e["student_id"]}|{e["halaqah_id"]}': (i + 1, e["total_baris"]) for i, e in enumerate(sorted_hafazan)}
    rank_murajaah_map = {f'{e["student_id"]}|{e["halaqah_id"]}': (i + 1, e["total_muka_surat"]) for i, e in enumerate(sorted_murajaah)}

    # Latest hafazan per (student_id, halaqah_id)
    latest_hafazan_map: Dict[str, Dict[str, Any]] = {}
    for r in sorted(records, key=lambda x: (x.get("tarikh") or "", x.get("created_at") or ""), reverse=True):
        if r.get("halaqah_id") != halaqah_id:
            continue
        if r.get("mod") != "hafazan_baru":
            continue
        if (r.get("tarikh") or "") > today_iso:
            continue
        sid = r["student_id"]
        if sid in latest_hafazan_map:
            continue
        latest_hafazan_map[sid] = {
            "juzuk_surah": r.get("juzuk_surah"),
            "muka_surat": r.get("muka_surat"),
            "tarikh": r.get("tarikh"),
        }

    # 4. Build summary
    out = []
    for s in all_students:
        sid = s["id"]
        my_key = f"{sid}|{halaqah_id}"
        prof = profile_map.get(sid)
        selected = prof.get("selected_juzuk_surah", []) if prof else []
        jumlah_juzuk = sum(1 for x in selected if x.startswith("Juzuk"))
        jumlah_surah = sum(1 for x in selected if not x.startswith("Juzuk"))
        rh = rank_hafazan_map.get(my_key)
        rm = rank_murajaah_map.get(my_key)
        out.append({
            "student_id": sid,
            "nama": s["nama"],
            "halaqah_id": halaqah_id,
            "source": s.get("source", "sheet"),
            "jumlah_juzuk": jumlah_juzuk,
            "jumlah_surah": jumlah_surah,
            "hafazan_terkini": latest_hafazan_map.get(sid),
            "ranking_hafazan": rh[0] if rh else None,
            "ranking_murajaah": rm[0] if rm else None,
            "total_baris": rh[1] if rh else 0,
            "total_muka_surat": rm[1] if rm else 0,
        })
    return out


@api_router.get("/student-profile/full")
async def get_student_profile_full(student_id: str, halaqah_id: str, _auth: bool = Depends(verify_token)):
    """Returns: selected_juzuk_surah, jumlah_juzuk, jumlah_surah,
    hafazan_terkini (latest hafazan_baru record), ranking_hafazan, ranking_murajaah."""
    key = {"student_id": student_id, "halaqah_id": halaqah_id}
    profile = await db.student_profiles.find_one(key, {"_id": 0})
    selected = profile.get("selected_juzuk_surah", []) if profile else []
    nama = profile.get("nama") if profile else None
    jumlah_juzuk = sum(1 for s in selected if s.startswith("Juzuk"))
    jumlah_surah = sum(1 for s in selected if not s.startswith("Juzuk"))

    # Latest hafazan_baru record (only up to today, to avoid future-dated records)
    today_iso = date.today().isoformat()
    latest = await db.tasmi_records.find_one(
        {
            "student_id": student_id,
            "halaqah_id": halaqah_id,
            "mod": "hafazan_baru",
            "tarikh": {"$lte": today_iso},
        },
        {"_id": 0},
        sort=[("tarikh", -1), ("created_at", -1)]
    )
    hafazan_terkini = None
    if latest:
        hafazan_terkini = {
            "juzuk_surah": latest.get("juzuk_surah"),
            "muka_surat": latest.get("muka_surat"),
            "tarikh": latest.get("tarikh"),
        }
        if not nama:
            nama = latest.get("student_nama")

    # Compute rankings from top-students analytics (overall, no month filter)
    records = await db.tasmi_records.find({}, {"_id": 0}).to_list(20000)
    hafazan_count: Dict[str, Dict[str, Any]] = {}
    murajaah_count: Dict[str, Dict[str, Any]] = {}
    for r in records:
        if not _is_countable(r):
            continue
        sid = r["student_id"]
        hid = r["halaqah_id"]
        k = f"{sid}|{hid}"
        if r["mod"] == "hafazan_baru":
            e = hafazan_count.setdefault(k, {"student_id": sid, "halaqah_id": hid, "total_baris": 0})
            e["total_baris"] += _safe_int(r.get("jumlah"))
        if r["mod"] == "murajaah":
            e = murajaah_count.setdefault(k, {"student_id": sid, "halaqah_id": hid, "total_muka_surat": 0})
            e["total_muka_surat"] += _safe_int(r.get("jumlah"))

    sorted_hafazan = sorted(hafazan_count.values(), key=lambda x: -x["total_baris"])
    sorted_murajaah = sorted(murajaah_count.values(), key=lambda x: -x["total_muka_surat"])

    my_key = f"{student_id}|{halaqah_id}"
    rank_hafazan = next((i + 1 for i, e in enumerate(sorted_hafazan) if f'{e["student_id"]}|{e["halaqah_id"]}' == my_key), None)
    rank_murajaah = next((i + 1 for i, e in enumerate(sorted_murajaah) if f'{e["student_id"]}|{e["halaqah_id"]}' == my_key), None)

    total_baris = next((e["total_baris"] for e in sorted_hafazan if f'{e["student_id"]}|{e["halaqah_id"]}' == my_key), 0)
    total_muka_surat = next((e["total_muka_surat"] for e in sorted_murajaah if f'{e["student_id"]}|{e["halaqah_id"]}' == my_key), 0)

    return {
        "student_id": student_id,
        "halaqah_id": halaqah_id,
        "nama": nama,
        "selected_juzuk_surah": selected,
        "jumlah_juzuk": jumlah_juzuk,
        "jumlah_surah": jumlah_surah,
        "hafazan_terkini": hafazan_terkini,
        "ranking_hafazan": rank_hafazan,
        "ranking_murajaah": rank_murajaah,
        "total_pelajar_hafazan": len(sorted_hafazan),
        "total_pelajar_murajaah": len(sorted_murajaah),
        "total_baris_hafazan": total_baris,
        "total_muka_surat_murajaah": total_muka_surat,
    }


@api_router.post("/admin/migrate-student-ids")
async def migrate_student_ids(_auth: bool = Depends(verify_teacher)):
    """One-off migration: update tasmi_records & student_profiles to use stable hash-based IDs."""
    import hashlib
    def stable_id(halaqah_id, nama):
        h = hashlib.md5((nama or "").upper().encode("utf-8")).hexdigest()[:8]
        return f"sheet_{halaqah_id}_{h}"

    updated_tasmi = 0
    updated_profiles = 0

    # Migrate tasmi_records (only sheet-based IDs, not db_xxxx)
    records = await db.tasmi_records.find({"student_id": {"$regex": "^sheet_"}}, {"_id": 0, "id": 1, "student_id": 1, "student_nama": 1, "halaqah_id": 1}).to_list(50000)
    for r in records:
        new_id = stable_id(r["halaqah_id"], r.get("student_nama"))
        if new_id != r["student_id"]:
            await db.tasmi_records.update_one({"id": r["id"]}, {"$set": {"student_id": new_id}})
            updated_tasmi += 1

    # Migrate student_profiles
    profiles = await db.student_profiles.find({"student_id": {"$regex": "^sheet_"}}, {"_id": 0}).to_list(10000)
    seen_keys = set()
    for p in profiles:
        new_id = stable_id(p["halaqah_id"], p.get("nama"))
        if new_id != p["student_id"]:
            new_key = f'{new_id}|{p["halaqah_id"]}'
            if new_key in seen_keys:
                # Duplicate after migration - delete old
                await db.student_profiles.delete_one({"student_id": p["student_id"], "halaqah_id": p["halaqah_id"]})
            else:
                # Check if target already exists
                existing = await db.student_profiles.find_one({"student_id": new_id, "halaqah_id": p["halaqah_id"]})
                if existing:
                    await db.student_profiles.delete_one({"student_id": p["student_id"], "halaqah_id": p["halaqah_id"]})
                else:
                    await db.student_profiles.update_one(
                        {"student_id": p["student_id"], "halaqah_id": p["halaqah_id"]},
                        {"$set": {"student_id": new_id}}
                    )
                seen_keys.add(new_key)
            updated_profiles += 1

    return {"updated_tasmi": updated_tasmi, "updated_profiles": updated_profiles}


@api_router.get("/")
async def root():
    return {"message": "e-HUFFAZ AL-AKHYAR API"}


# ---------- Backup & Export ----------
MODE_LABELS_MS = {
    "hafazan_baru": "Hafazan Baru",
    "murajaah": "Murajaah",
    "tilawah": "Tilawah",
    "iqra": "Iqra'",
    "amoktha_khatam": "Amoktha Khatam",
}
KEPUTUSAN_LABELS_MS = {
    "mumtaz": "Mumtaz",
    "jayyid": "Jayyid",
    "daif": "Daif",
    "gagal_hantar": "Gagal Hantar",
    "tidak_hadir": "Tidak Hadir",
}
HALAQAH_NAME_BY_ID = {h["id"]: h["name"] for h in HALAQAHS}


async def _fetch_backup_data():
    buletin = await db.buletin.find({}, {"_id": 0}).sort("created_at", -1).to_list(100000)
    tasmi = await db.tasmi_records.find({}, {"_id": 0}).sort("tarikh", -1).to_list(100000)
    extra = await db.extra_students.find({}, {"_id": 0}).to_list(100000)
    # Enrich
    for t in tasmi:
        t["halaqah_nama"] = HALAQAH_NAME_BY_ID.get(t.get("halaqah_id"), t.get("halaqah_id"))
        t["mod_label"] = MODE_LABELS_MS.get(t.get("mod"), t.get("mod"))
        t["keputusan_label"] = KEPUTUSAN_LABELS_MS.get(t.get("keputusan"), t.get("keputusan") or "")
    for e in extra:
        e["halaqah_nama"] = HALAQAH_NAME_BY_ID.get(e.get("halaqah_id"), e.get("halaqah_id"))
    return buletin, tasmi, extra


def _rows_to_csv(headers, rows, row_picker):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for r in rows:
        writer.writerow(row_picker(r))
    return buf.getvalue()


BULETIN_HEADERS = ["ID", "Tajuk", "Memo", "Penghantar", "Tarikh", "Dicipta Pada"]
TASMI_HEADERS = ["ID", "Tarikh", "Tarikh Akhir", "Jenis", "Halaqah", "Pelajar", "Mod Tasmi'", "Juzuk/Surah", "IQRA", "Muka Surat", "Jumlah", "Jenis Jumlah", "Keputusan", "Catatan", "Dicipta Pada"]
EXTRA_HEADERS = ["ID", "Nama Pelajar", "Halaqah", "Sumber"]


def _buletin_row(r):
    return [r.get("id"), r.get("tajuk"), r.get("memo"), r.get("penghantar"), r.get("tarikh"), r.get("created_at")]


def _tasmi_row(r):
    return [
        r.get("id"), r.get("tarikh"), r.get("tarikh_akhir") or "",
        "Mingguan" if r.get("is_weekly") else "Harian",
        r.get("halaqah_nama"), r.get("student_nama"),
        r.get("mod_label"), r.get("juzuk_surah") or "", r.get("iqra_level") or "",
        r.get("muka_surat") or "", r.get("jumlah") or "", r.get("jumlah_type") or "",
        r.get("keputusan_label"), r.get("catatan") or "", r.get("created_at"),
    ]


def _extra_row(r):
    return [r.get("id"), r.get("nama"), r.get("halaqah_nama"), r.get("source")]


@api_router.get("/backup/zip")
async def backup_zip(_auth: bool = Depends(verify_teacher)):
    buletin, tasmi, extra = await _fetch_backup_data()
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("buletin.csv", _rows_to_csv(BULETIN_HEADERS, buletin, _buletin_row))
        zf.writestr("tasmi_records.csv", _rows_to_csv(TASMI_HEADERS, tasmi, _tasmi_row))
        zf.writestr("extra_students.csv", _rows_to_csv(EXTRA_HEADERS, extra, _extra_row))
    buf.seek(0)
    today = date.today().isoformat()
    filename = f"e-huffaz-backup-{today}.zip"
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@api_router.get("/backup/excel")
async def backup_excel(_auth: bool = Depends(verify_teacher)):
    buletin, tasmi, extra = await _fetch_backup_data()
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="166534")
    center = Alignment(horizontal="center", vertical="center")

    def write_sheet(ws, headers, rows, row_picker):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        for r in rows:
            ws.append(row_picker(r))
        # Auto column width
        for i, h in enumerate(headers, start=1):
            max_len = len(str(h))
            for r in rows:
                v = row_picker(r)[i - 1]
                if v and len(str(v)) > max_len:
                    max_len = min(len(str(v)), 60)
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max_len + 2

    ws1 = wb.active
    ws1.title = "Buletin"
    write_sheet(ws1, BULETIN_HEADERS, buletin, _buletin_row)

    ws2 = wb.create_sheet("Rekod Tasmi")
    write_sheet(ws2, TASMI_HEADERS, tasmi, _tasmi_row)

    ws3 = wb.create_sheet("Pelajar Tambahan")
    write_sheet(ws3, EXTRA_HEADERS, extra, _extra_row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    today = date.today().isoformat()
    filename = f"e-huffaz-backup-{today}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@api_router.get("/backup/summary")
async def backup_summary(_auth: bool = Depends(verify_teacher)):
    counts = await asyncio_gather_counts()
    return counts


async def asyncio_gather_counts():
    buletin_count = await db.buletin.count_documents({})
    tasmi_count = await db.tasmi_records.count_documents({})
    extra_count = await db.extra_students.count_documents({})
    return {
        "buletin": buletin_count,
        "tasmi_records": tasmi_count,
        "extra_students": extra_count,
    }


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
